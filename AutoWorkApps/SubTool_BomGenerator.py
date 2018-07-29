# -*- coding: utf-8 -*-
"""
Created on Wed Jul  4 20:27:52 2018

@author: Dezter
"""
def ST_BomGenerator():
    import time
    import openpyxl as excel
    import os
    
    TimeStart = time.time()
    """ BOM Infomation """
    QPN_31 = "41PCBA,51PCBA"
    QPN_41 = "41G7CCS0120"
    QPN_51 = "51G7CSS0120"
    QPN_PCB = "DAG7CDMB6A0"
    Head = ["Parent","Part Number","Item Number","Alt Grp","Usage(%)","Qty","Location","Function Code"]
    PDM_bot_Head = [QPN_41,QPN_51,"10","","",1,"",""]
    PDM_top_Head = [QPN_51,QPN_PCB,"10","","",1,"",""]
    """ BOM File loading """
    BomFile = "D:/Python/ForWork/BomGenerator/BOM/G7C120.BOM"
    Bom = open(BomFile,"r").readlines()
    BomA = [Bom[i].strip() for i in range(len(Bom))]
    BomList = [BomA[i].split("\t") for i in range(len(BomA))]
    
    Com_botFile = "D:/Python/ForWork/BomGenerator/BOM/cmp_bot.txt"
    Com_bot = open(Com_botFile,"r").readlines()
    Com_bot = [Com_bot[i].split( ) for i in range(len(Com_bot))]
    
    Com_topFile = "D:/Python/ForWork/BomGenerator/BOM/cmp_top.txt"
    Com_top = open(Com_topFile,"r").readlines()
    Com_top = [Com_top[i].split( ) for i in range(len(Com_top))]
    """ BOM Generate """
    BomListA = BomList[14:]
    Com_botA = [Com_bot[i] for i in range(len(Com_bot)) if len(Com_bot[i]) == 9]
    Com_topA = [Com_top[i] for i in range(len(Com_top)) if len(Com_top[i]) == 9]
    BomListAT = list(map(list, zip(*BomListA)))
    #Com_butAT = list(map(list, zip(*Com_butA)))
    #Com_topAT = list(map(list, zip(*Com_topA)))
    #--- Separate BOM to Top & button part ---
    BomBot = []
    BomTop = []
    for i in range(len(Com_botA)):
        BomBot.append(BomListA[BomListAT[0].index(Com_botA[i][0])])
    for i in range(len(Com_topA)):
        BomTop.append(BomListA[BomListAT[0].index(Com_topA[i][0])])
    
    #--- pick up usage materials from Top & button part ---
    BomBotA = [BomBot[i] for i in range(len(BomBot)) if BomBot[i][1].find("*") < 0]
    BomTopA = [BomTop[i] for i in range(len(BomTop)) if BomTop[i][1].find("*") < 0]
    
    #--- Sort the List by QPN for integating into formal BOM format ---
    BomBotB = sorted(BomBotA, key = lambda a: a[2])
    BomTopB = sorted(BomTopA, key = lambda a: a[2])
    BomBotBT = list(map(list, zip(*BomBotB)))
    BomTopBT = list(map(list, zip(*BomTopB)))
    BomBotbyQPN = list(set(BomBotBT[2]))
    BomTopbyQPN = list(set(BomTopBT[2]))
    BomBotbyQPN = sorted(BomBotbyQPN, key = lambda a:a)
    BomTopbyQPN = sorted(BomTopbyQPN, key = lambda a:a)
    #--- BOM table format parameter---
    PDM_bot = []
    PDM_top = []
    Col_Parent = 0
    Col_PN = 1
    Col_Item = 2
    Col_Grp = 3
    Col_Usage = 4
    Col_Qty = 5
    Col_Loca = 6
    Col_FunCode = 7
    #--- integating BOM from row data by Location to row data by QPN ---
    for i in range(len(BomBotbyQPN)):
        a = BomBotbyQPN[i]
        Alist = []
        Alist = BomBotBT[0][BomBotBT[2].index(a):BomBotBT[2].index(a)+BomBotBT[2].count(a)]
        Avalue = ""
        for i1 in Alist:
            Avalue = Avalue + i1 + ","
        Avalue = Avalue.rstrip(",")
        PDM_bot_data = [QPN_41,a,"","","",BomBotBT[2].count(a),Avalue,BomBotBT[4][BomBotBT[2].index(a):BomBotBT[2].index(a)+BomBotBT[2].count(a)]]
        
        PDM_bot.append(PDM_bot_data)
        
    for j in range(len(BomTopbyQPN)):
        b = BomTopbyQPN[j]
        Blist = []
        Blist = BomTopBT[0][BomTopBT[2].index(b):BomTopBT[2].index(b)+BomTopBT[2].count(b)]
        Bvalue = ""
        for j1 in Blist:
            Bvalue = Bvalue + j1 + ","
        Bvalue = Bvalue.rstrip(",")    
        PDM_top_data = [QPN_51,b,"","","",BomTopBT[2].count(b),Bvalue,BomTopBT[4][BomTopBT[2].index(b):BomTopBT[2].index(b)+BomTopBT[2].count(b)]]
        PDM_top.append(PDM_top_data)
        
    #--- Simplify Function Code from raw data ---
    for i in range(len(PDM_bot)):
        i_Set = list(set(PDM_bot[i][Col_FunCode]))
        i_Sort = sorted(i_Set)
        Alist = []
        Alist = [i_Sort[n]+str("{:0>2d}".format(PDM_bot[i][Col_FunCode].count(i_Sort[n]))) for n in range(len(i_Sort))]
        Avalue = ""
        for i1 in Alist:
            Avalue = Avalue + i1 + ","
        Avalue = Avalue.rstrip(",")
        PDM_bot[i][Col_FunCode] = Avalue
    
    for j in range(len(PDM_top)):
        j_Set = list(set(PDM_top[j][Col_FunCode]))
        j_Sort = sorted(j_Set)
        Blist = []
        Blist = [j_Sort[n]+str("{:0>2d}".format(PDM_top[j][Col_FunCode].count(j_Sort[n]))) for n in range(len(j_Sort))]
        Bvalue = ""
        for j1 in Blist:
            Bvalue = Bvalue + j1 + ","
        Bvalue = Bvalue.rstrip(",")
        PDM_top[j][Col_FunCode] = Bvalue
    
    #--- arrange information from downBOM ---
    PDM_botT = list(map(list, zip(*PDM_bot)))
    PDM_topT = list(map(list, zip(*PDM_top)))
    downBOM = "D:/Python/ForWork/BomGenerator/BOM/downbom_G7C120.xlsx"
    if os.path.exists(downBOM) is False:
        pass
        """
        PDM_botT[Col_Item] = [(n+2)*10 for n in range(len(PDM_botT[Col_Item]))]
        PDM_topT[Col_Item] = [(n+2)*10 for n in range(len(PDM_topT[Col_Item]))]
        PDM_bot = list(map(list, zip(*PDM_botT)))
        PDM_top = list(map(list, zip(*PDM_topT)))
        """
    elif os.path.exists(downBOM) is True:
        wb = excel.load_workbook(downBOM)
        ws41 = wb["41"]
        rows = ws41.rows
        downBOM41 = []
        for row in rows:
            line = [col.value for col in row]
            downBOM41.append(line)
        ws51 = wb["51"]
        rows = ws51.rows
        downBOM51 = []
        for row in rows:
            line = [col.value for col in row]
            downBOM51.append(line)
        #--- pick out the useful data ---    
        downBOM41_QPN = downBOM41[0].index("Part Number")
        downBOM41_Item = downBOM41[0].index("Item Number")
        downBOM41_Group = downBOM41[0].index("Alt Grp")
        downBOM41_Usage = downBOM41[0].index("Usage(%)")
        downBOM51_QPN = downBOM51[0].index("Part Number")
        downBOM51_Item = downBOM51[0].index("Item Number")
        downBOM51_Group = downBOM51[0].index("Alt Grp")
        downBOM51_Usage = downBOM51[0].index("Usage(%)")
        downBOM41 = downBOM41[1:]
        downBOM51 = downBOM51[1:]
        downBOM41_QIGU = [[downBOM41[i][downBOM41_QPN],downBOM41[i][downBOM41_Item],downBOM41[i][downBOM41_Group],downBOM41[i][downBOM41_Usage]] for i in range(len(downBOM41)) if downBOM41[i][downBOM41_QPN] is not None]
        downBOM51_QIGU = [[downBOM51[i][downBOM51_QPN],downBOM51[i][downBOM51_Item],downBOM51[i][downBOM41_Group],downBOM51[i][downBOM41_Usage]] for i in range(len(downBOM51)) if downBOM51[i][downBOM51_QPN] is not None]
        downBOM41_QIGU = sorted(downBOM41_QIGU, key = lambda a: a[0])
        downBOM51_QIGU = sorted(downBOM51_QIGU, key = lambda a: a[0])
        downBOM41T_QIGU = list(map(list, zip(*downBOM41_QIGU)))
        downBOM51T_QIGU = list(map(list, zip(*downBOM51_QIGU)))
        #--- mapping the item number from downBOM to BOM ---
        """
        PDM_bot_I = [PDM_bot[i] for i in range(len(PDM_bot)) if PDM_bot[i][1] in downBOM41T_QIGU[0]]
        PDM_bot_D = [PDM_bot[i] for i in range(len(PDM_bot)) if PDM_bot[i][1] not in downBOM41T_QIGU[0]]
        PDM_top_I = [PDM_top[i] for i in range(len(PDM_top)) if PDM_top[i][1] in downBOM51T_QIGU[0]]
        PDM_top_D = [PDM_top[i] for i in range(len(PDM_top)) if PDM_top[i][1] not in downBOM51T_QIGU[0]]    
        downBOM41_QIGU_I = [downBOM41_QIGU[i] for i in range(len(downBOM41_QIGU)) if downBOM41_QIGU[i][0] in PDM_botT[1]]
        downBOM51_QIGU_I = [downBOM51_QIGU[i] for i in range(len(downBOM51_QIGU)) if downBOM51_QIGU[i][0] in PDM_topT[1]]
        PDM_botT_I = list(map(list, zip(*PDM_bot_I)))
        PDM_topT_I = list(map(list, zip(*PDM_top_I)))
        downBOM41T_QIGU_I = list(map(list, zip(*downBOM41_QIGU_I)))
        downBOM51T_QIGU_I = list(map(list, zip(*downBOM51_QIGU_I)))
        PDM_botT_I[Col_Item] = downBOM41T_QIGU_I[1]
        PDM_topT_I[Col_Item] = downBOM51T_QIGU_I[1]
        PDM_bot_I = list(map(list, zip(*PDM_botT_I)))
        PDM_top_I = list(map(list, zip(*PDM_topT_I)))
        #--- fill the item number in new components ---
        PDM_botT_D = list(map(list, zip(*PDM_bot_D)))
        PDM_topT_D = list(map(list, zip(*PDM_top_D)))
        PDM_botT_D[Col_Item] = [max(i for i in downBOM41T_QIGU[1] if type(i) is int)+(n+1)*10 for n in range(len(PDM_botT_D[Col_Item]))]
        PDM_topT_D[Col_Item] = [max(i for i in downBOM51T_QIGU[1] if type(i) is int)+(n+1)*10 for n in range(len(PDM_topT_D[Col_Item]))]
        #PDM_botT_D[Col_Item] = [(n+2)*10 for n in range(len(PDM_botT[Col_Item])) if (n+2)*10 not in downBOM41T_QI_I[1]]
        #PDM_topT_D[Col_Item] = [(n+2)*10 for n in range(len(PDM_topT[Col_Item])) if (n+2)*10 not in downBOM51T_QI_I[1]]
        PDM_bot_D = list(map(list, zip(*PDM_botT_D)))
        PDM_top_D = list(map(list, zip(*PDM_topT_D)))
        PDM_bot = PDM_bot_I + PDM_bot_D
        PDM_top = PDM_top_I + PDM_top_D
        PDM_bot = sorted(PDM_bot, key = lambda a: a[1])
        PDM_top = sorted(PDM_top, key = lambda a: a[1])
        """
    PDM_bot.insert(0,PDM_bot_Head)
    PDM_top.insert(0,PDM_top_Head)
    PDM_botT = list(map(list, zip(*PDM_bot)))
    PDM_topT = list(map(list, zip(*PDM_top)))
    
    #--- Add 2nd source or pilot run 2nd source ---
    File_Modify =  "D:/Python/ForWork/BomGenerator/BOM/Modify_List.xlsx"
    if os.path.exists(File_Modify) is True:
        wb = excel.load_workbook(File_Modify)
        sheets = wb.get_sheet_names()
        temp = wb.get_sheet_by_name("Change") 
        rows = temp.rows
        ModifyList = []
        for row in rows:
            line = [col.value for col in row]
            ModifyList.append(line)
            
        ModifyListT = list(map(list, zip(*ModifyList)))
        ColLen = len([i for i in ModifyListT[1] if i is not None])
        RowLen = max([len([i for i in ModifyList[j] if i is not None]) for j in range(ColLen+1)])
        ModifyList = [ModifyList[i][0:RowLen] for i in range(ColLen+1)]
        #--- Seprate different use from Modify List --- 
        ModifyHead = ModifyList[0][2:]
        Modify = ModifyList[2:]
        ModifyT = list(map(list, zip(*Modify)))
        Modify_Mate_Add = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is None and ModifyT[1][i] == "add"]
        Modify_Mate_Change = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is None and ModifyT[1][i] == "change"]
        Modify_Mate_Del = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is None and ModifyT[1][i] == "del"]
        Modify_Loca_Add = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is not None and ModifyT[1][i] == "add"]
        Modify_Loca_Change = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is not None and ModifyT[1][i] == "change"]
        Modify_Loca_Del = [Modify[i] for i in range(len(Modify)) if ModifyT[0][i] is not None and ModifyT[1][i] == "del"]
        Modify_Mate_AddT = list(map(list, zip(*Modify_Mate_Add)))
        Modify_Mate_ChangeT = list(map(list, zip(*Modify_Mate_Change)))
        Modify_Mate_DelT = list(map(list, zip(*Modify_Mate_Del)))
        Modify_Loca_AddT = list(map(list, zip(*Modify_Loca_Add)))
        Modify_Loca_ChangeT = list(map(list, zip(*Modify_Loca_Change)))
        Modify_Loca_DelT = list(map(list, zip(*Modify_Loca_Del)))
        #--- For Add Material ---     
        if Modify_Mate_Add == []:
            pass
        else:
            Group_Mate_Add = []
            List_Mate_Add = []
            Usage_Mat_Add = []
            
            #--- Transfer matrix data to row data and mark the group number to help us identify row data, instead of multi-for loop ---
            
            for i in range(len(Modify_Mate_Add)):
                Group_Mate_Add = Group_Mate_Add+[str(i)]*len(list(filter(None,Modify_Mate_Add[i][2:])))
                List_Mate_Add = List_Mate_Add+list(filter(None,Modify_Mate_Add[i][2:]))
                Usage_Mat_Add = Usage_Mat_Add+[100]+[0]*(len(list(filter(None,Modify_Mate_Add[i][2:])))-1)
            GroupwListT = [Group_Mate_Add,List_Mate_Add,Usage_Mat_Add]
            GroupwList = list(map(list, zip(*GroupwListT)))                
            #--- Pick out Topside's 2nd Source and fill in group & usage by rule ---
            List_Mate_Add_top = list(set(List_Mate_Add) & (set(PDM_topT[1])))
            SecondRow = []
            SecondQPN = []
            List_Mate_Add_topG = []
            for j in List_Mate_Add_top:
                Group = GroupwListT[0][GroupwListT[1].index(j)]
                Count = GroupwListT[0].count(Group)
                Second = [PDM_top[PDM_topT[Col_PN].index(j)].copy()]*(Count)
                SecondRow = SecondRow + Second
                SecondQPN = SecondQPN + GroupwListT[1][GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
                del(PDM_top[PDM_topT[Col_PN].index(j)])
                PDM_topT[Col_PN].remove(j)
                PDM_topT = list(map(list, zip(*PDM_top)))
                List_Mate_Add_topG = List_Mate_Add_topG + GroupwList[GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
            
            List_Mate_Add_topGT = list(map(list, zip(*List_Mate_Add_topG)))
            for k in List_Mate_Add_topGT[1]:
                if 'downBOM51T_QIGU' in dir() and k in downBOM51T_QIGU[0] and downBOM51_QIGU[downBOM51T_QIGU[0].index(k)][-2] is not None:
                    List_Mate_Add_topG[List_Mate_Add_topGT[1].index(k)].extend(downBOM51_QIGU[downBOM51T_QIGU[0].index(k)][-2:])
                else:
                    List_Mate_Add_topG[List_Mate_Add_topGT[1].index(k)].extend(["",""])
            GroupwList = list(map(list, zip(*GroupwListT)))        
            if 'downBOM51T_QIGU' in dir():
                GMark_top = list(max(max(x for x in downBOM51T_QIGU[2] if x is not None),max(y for y in PDM_topT[Col_Grp])))
            else:
                GMark_top = list(max(max(y for y in PDM_topT[Col_Grp]),"A0"))            
            WtN = ord(GMark_top[0])*10+int(GMark_top[1])
            List_Mate_Add_topG = sorted(List_Mate_Add_topG ,key = lambda a:(a[0],a[-1],a[2]),reverse=True)
            List_Mate_Add_topGT = list(map(list, zip(*List_Mate_Add_topG)))
            for m in list(set(List_Mate_Add_topGT[0])):
                Group = List_Mate_Add_topGT[0].index(m)
                Count = List_Mate_Add_topGT[0].count(m)
                if List_Mate_Add_topGT[-1][Group] == 100:
                    List_Mate_Add_topGT[-2][Group:Group+Count]=[List_Mate_Add_topGT[-2][Group]]*len(List_Mate_Add_topGT[-2][Group:Group+Count])
                    List_Mate_Add_topGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Add_topGT[-1][Group+1:Group+Count])
                elif List_Mate_Add_topGT[-1][Group] == 0:
                    List_Mate_Add_topGT[-2][Group:Group+Count]=[List_Mate_Add_topGT[-2][Group]]*len(List_Mate_Add_topGT[-2][Group:Group+Count])
                    List_Mate_Add_topGT[-1][Group+1]=100
                    List_Mate_Add_topGT[-1][Group+2:Group+Count]=[0]*len(List_Mate_Add_topGT[-1][Group+2:Group+Count])
                elif List_Mate_Add_topGT[-1][Group] == "":
                    WtN = WtN+1
                    if WtN%10 == 0:
                        WtN = WtN + 1
                    else:
                        pass
                    NtW = chr(WtN//10)+str(WtN%10)
                    List_Mate_Add_topGT[-2][Group:Group+Count]=[NtW]*len(List_Mate_Add_topGT[-2][Group:Group+Count])
                    List_Mate_Add_topGT[-1][Group]=100
                    List_Mate_Add_topGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Add_topGT[-1][Group+1:Group+Count])             
            List_Mate_Add_topG = list(map(list, zip(*List_Mate_Add_topGT)))    
            List_Mate_Add_topG = [List_Mate_Add_topG[List_Mate_Add_topGT[1].index(i)] for i in SecondQPN if i in List_Mate_Add_topGT[1]]
            SecondRowT = list(map(list, zip(*SecondRow)))
            List_Mate_Add_topGT = list(map(list, zip(*List_Mate_Add_topG)))
            SecondRowT[1] = SecondQPN
            SecondRowT[3] = List_Mate_Add_topGT[3] 
            SecondRowT[4] = List_Mate_Add_topGT[4] 
            SecondRow = list(map(list, zip(*SecondRowT)))
            PDM_top = PDM_top + SecondRow
            PDM_top = sorted(PDM_top, key = lambda a: (a[Col_Item],a[Col_PN]))
            PDM_topT = list(map(list, zip(*PDM_top)))
            
            #--- Pick out Botside's 2nd Source and fill in group & usage by rule ---
    
            List_Mate_Add_bot = list(set(List_Mate_Add) & (set(PDM_botT[1])))
            SecondRow = []
            SecondQPN = []
            List_Mate_Add_botG = []
            for j in List_Mate_Add_bot:
                Group = GroupwListT[0][GroupwListT[1].index(j)]
                Count = GroupwListT[0].count(Group)
                Second = [PDM_bot[PDM_botT[Col_PN].index(j)].copy()]*(Count)
                SecondRow = SecondRow + Second
                SecondQPN = SecondQPN + GroupwListT[1][GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
                del(PDM_bot[PDM_botT[Col_PN].index(j)])
                PDM_botT[Col_PN].remove(j)
                List_Mate_Add_botG = List_Mate_Add_botG + GroupwList[GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
            
            List_Mate_Add_botGT = list(map(list, zip(*List_Mate_Add_botG)))
            for k in List_Mate_Add_botGT[1]:
                if 'downBOM41T_QIGU' in dir() and k in downBOM41T_QIGU[0] and downBOM41_QIGU[downBOM41T_QIGU[0].index(k)][-2] is not None:
                    List_Mate_Add_botG[List_Mate_Add_botGT[1].index(k)].extend(downBOM41_QIGU[downBOM41T_QIGU[0].index(k)][-2:])
                else:
                    List_Mate_Add_botG[List_Mate_Add_botGT[1].index(k)].extend(["",""])
            GroupwList = list(map(list, zip(*GroupwListT)))
            
            if 'downBOM41T_QIGU' in dir():
                GMark_bot = list(max(max(x for x in downBOM41T_QIGU[2] if x is not None),max(y for y in PDM_botT[Col_Grp])))
            else:
                GMark_bot = list(max(max(y for y in PDM_botT[Col_Grp]),"A0"))
            WtN = ord(GMark_bot[0])*10+int(GMark_bot[1])
            List_Mate_Add_botG = sorted(List_Mate_Add_botG ,key = lambda a:(a[0],a[-1],a[2]),reverse=True)
            List_Mate_Add_botGT = list(map(list, zip(*List_Mate_Add_botG)))
            for m in list(set(List_Mate_Add_botGT[0])):
                Group = List_Mate_Add_botGT[0].index(m)
                Count = List_Mate_Add_botGT[0].count(m)
                if List_Mate_Add_botGT[-1][Group] == 100:
                    List_Mate_Add_botGT[-2][Group:Group+Count]=[List_Mate_Add_botGT[-2][Group]]*len(List_Mate_Add_botGT[-2][Group:Group+Count])
                    List_Mate_Add_botGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Add_botGT[-1][Group+1:Group+Count])
                elif List_Mate_Add_botGT[-1][Group] == 0:
                    List_Mate_Add_botGT[-2][Group:Group+Count]=[List_Mate_Add_botGT[-2][Group]]*len(List_Mate_Add_botGT[-2][Group:Group+Count])
                    List_Mate_Add_botGT[-1][Group+1]=100
                    List_Mate_Add_botGT[-1][Group+2:Group+Count]=[0]*len(List_Mate_Add_botGT[-1][Group+2:Group+Count])
                elif List_Mate_Add_botGT[-1][Group] == "":
                    WtN = WtN+1
                    if WtN%10 == 0:
                        WtN = WtN + 1
                    else:
                        pass
                    NtW = chr(WtN//10)+str(WtN%10)
                    List_Mate_Add_botGT[-2][Group:Group+Count]=[NtW]*len(List_Mate_Add_botGT[-2][Group:Group+Count])
                    List_Mate_Add_botGT[-1][Group]=100
                    List_Mate_Add_botGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Add_botGT[-1][Group+1:Group+Count])             
            List_Mate_Add_botG = list(map(list, zip(*List_Mate_Add_botGT)))    
            List_Mate_Add_botG = [List_Mate_Add_botG[List_Mate_Add_botGT[1].index(i)] for i in SecondQPN if i in List_Mate_Add_botGT[1]]
            SecondRowT = list(map(list, zip(*SecondRow)))
            List_Mate_Add_botGT = list(map(list, zip(*List_Mate_Add_botG)))
            SecondRowT[1] = SecondQPN
            SecondRowT[3] = List_Mate_Add_botGT[3] 
            SecondRowT[4] = List_Mate_Add_botGT[4] 
            SecondRow = list(map(list, zip(*SecondRowT)))
            PDM_bot = PDM_bot + SecondRow
            PDM_bot = sorted(PDM_bot, key = lambda a: (a[Col_Item],a[Col_PN]))
            PDM_botT = list(map(list, zip(*PDM_bot)))
        
        #--- For Del Material ---
        if Modify_Mate_Del == []:
            pass
        else:
            Group_Mate_Del = []
            List_Mate_Del = []
            Usage_Mat_Del = []
            #--- Transfer matrix data to row data and mark the group number to help us identify row data, instead of multi-for loop ---
            for i in range(len(Modify_Mate_Del)):
                Group_Mate_Del = Group_Mate_Del+[str(i)]*len(list(filter(None,Modify_Mate_Del[i][2:])))
                List_Mate_Del = List_Mate_Del+list(filter(None,Modify_Mate_Del[i][2:]))
                Usage_Mat_Del = Usage_Mat_Del+[100]+[0]*(len(list(filter(None,Modify_Mate_Del[i][2:])))-1)
            #GroupwListT = [Group_Mate_Del,List_Mate_Del,Usage_Mat_Del]
            #GroupwList = list(map(list, zip(*GroupwListT)))                
            #--- Pick out Topside's 2nd Source and fill in group & usage by rule ---
            List_Mate_Del_top = list(set(List_Mate_Del) & (set(PDM_topT[1])))
            for j in List_Mate_Del_top:
                del(PDM_top[PDM_topT[Col_PN].index(j)])
                PDM_topT[Col_PN].remove(j)
            PDM_topT = list(map(list, zip(*PDM_top)))        
            #--- Pick out Botside's 2nd Source and fill in group & usage by rule ---
            List_Mate_Del_bot = list(set(List_Mate_Del) & (set(PDM_botT[1])))
            for j in List_Mate_Del_bot:
                del(PDM_bot[PDM_botT[Col_PN].index(j)])
                PDM_botT[Col_PN].remove(j)
            PDM_botT = list(map(list, zip(*PDM_bot)))
        
        #--- For Change Material ---
        Group_ModifyHead = list(set(ModifyHead))
        
        if Modify_Mate_Change == []:
            pass
        elif Group_ModifyHead != [None] and QPN_31 not in Group_ModifyHead:
            pass
        else:
            pass
    
            List_Mate_ChangeAll = []
            Group_Mate_ChangeAll = []
            Group_Mate_Change = []
            List_Mate_Change = []
            Usage_Mat_Change = []        
            #--- Transfer matrix data to row data and mark the group number to help us identify row data, instead of multi-for loop ---
            for i in range(len(Modify_Mate_Change)):            
                List_Mate_ChangeAll = List_Mate_ChangeAll+list(filter(None,Modify_Mate_Change[i][2:]))
                Group_Mate_ChangeAll = Group_Mate_ChangeAll+[str(i)]*len(list(filter(None,Modify_Mate_Change[i][2:])))
                if Group_ModifyHead == [None]:
                    List_Mate_Change = List_Mate_Change+list(filter(None,Modify_Mate_Change[i][3:]))
                    Group_Mate_Change = Group_Mate_Change+[str(i)]*len(list(filter(None,Modify_Mate_Change[i][3:])))
                    if len(list(filter(None,Modify_Mate_Change[i][3:]))) == 1:
                        Usage_Mat_Change = Usage_Mat_Change+[""]
                    else:
                        Usage_Mat_Change = Usage_Mat_Change+[100]+[0]*(len(list(filter(None,Modify_Mate_Change[i][3:])))-1)
                    
                elif QPN_31 in Group_ModifyHead:
                    # how to use key woro to lock the postion instead of all words
                    List_Mate_Change = List_Mate_Change+list(filter(None,Modify_Mate_Change[i][3:]))
                    Group_Mate_Change = Group_Mate_Change+[str(i)]*len(list(filter(None,Modify_Mate_Change[i][3:])))
                    if len(list(filter(None,Modify_Mate_Change[i][3:]))) == 1:
                        Usage_Mat_Change = Usage_Mat_Change+[""]
                    else:
                        Usage_Mat_Change = Usage_Mat_Change+[100]+[0]*(len(list(filter(None,Modify_Mate_Change[i][3:])))-1)
            
            GroupwListAllT = [Group_Mate_ChangeAll,List_Mate_ChangeAll]
            GroupwListAll = list(map(list, zip(*GroupwListAllT)))        
            GroupwListT = [Group_Mate_Change,List_Mate_Change,Usage_Mat_Change]
            GroupwList = list(map(list, zip(*GroupwListT))) 
                 
            #--- Pick out Topside's 2nd Source and fill in group & usage by rule ---        
            List_Mate_ChangeAll_top = list(set(List_Mate_ChangeAll) & (set(PDM_topT[1])))
            SecondRow = []
            SecondQPN = []
            List_Mate_Change_topG = []
            for j in List_Mate_ChangeAll_top:
                Group = GroupwListAllT[0][GroupwListAllT[1].index(j)]
                Count = GroupwListT[0].count(Group)
                Second = [PDM_top[PDM_topT[Col_PN].index(j)].copy()]*(Count)
                SecondRow = SecondRow + Second
                SecondQPN = SecondQPN + GroupwListT[1][GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
                del(PDM_top[PDM_topT[Col_PN].index(j)])
                PDM_topT[Col_PN].remove(j)
                List_Mate_Change_topG = List_Mate_Change_topG + GroupwList[GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
            
            List_Mate_Change_topGT = list(map(list, zip(*List_Mate_Change_topG)))
            for k in List_Mate_Change_topGT[1]:
                if 'downBOM51T_QIGU' in dir() and k in downBOM51T_QIGU[0] and downBOM51_QIGU[downBOM51T_QIGU[0].index(k)][-2] is not None:
                    List_Mate_Change_topG[List_Mate_Change_topGT[1].index(k)].extend(downBOM51_QIGU[downBOM51T_QIGU[0].index(k)][-2:])
                else:
                    List_Mate_Change_topG[List_Mate_Change_topGT[1].index(k)].extend(["",""])
            GroupwList = list(map(list, zip(*GroupwListT)))        
            if 'downBOM51T_QIGU' in dir():
                GMark_top = list(max(max(x for x in downBOM51T_QIGU[2] if x is not None),max(y for y in PDM_topT[Col_Grp])))
            else:
                GMark_top = list(max(max(y for y in PDM_topT[Col_Grp]),"A0"))
            WtN = ord(GMark_top[0])*10+int(GMark_top[1])
            List_Mate_Change_topG = sorted(List_Mate_Change_topG ,key = lambda a:(a[0],a[-1],a[2]),reverse=True)
            List_Mate_Change_topGT = list(map(list, zip(*List_Mate_Change_topG)))
            for m in list(set(List_Mate_Change_topGT[0])):
                Group = List_Mate_Change_topGT[0].index(m)
                Count = List_Mate_Change_topGT[0].count(m)
                if Count == 1:
                    pass
                elif List_Mate_Change_topGT[-1][Group] == 100:
                    List_Mate_Change_topGT[-2][Group:Group+Count]=[List_Mate_Change_topGT[-2][Group]]*len(List_Mate_Change_topGT[-2][Group:Group+Count])
                    List_Mate_Change_topGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Change_topGT[-1][Group+1:Group+Count])
                elif List_Mate_Change_topGT[-1][Group] == 0:
                    List_Mate_Change_topGT[-2][Group:Group+Count]=[List_Mate_Add_topGT[-2][Group]]*len(List_Mate_Add_topGT[-2][Group:Group+Count])
                    List_Mate_Change_topGT[-1][Group+1]=100
                    List_Mate_Change_topGT[-1][Group+2:Group+Count]=[0]*len(List_Mate_Change_topGT[-1][Group+2:Group+Count])
                elif List_Mate_Change_topGT[-1][Group] == "":
                    WtN = WtN+1
                    if WtN%10 == 0:
                        WtN = WtN + 1
                    else:
                        pass
                    NtW = chr(WtN//10)+str(WtN%10)
                    List_Mate_Change_topGT[-2][Group:Group+Count]=[NtW]*len(List_Mate_Change_topGT[-2][Group:Group+Count])
                    List_Mate_Change_topGT[-1][Group]=100
                    List_Mate_Change_topGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Change_topGT[-1][Group+1:Group+Count])             
            List_Mate_Change_topG = list(map(list, zip(*List_Mate_Change_topGT)))    
            List_Mate_Change_topG = [List_Mate_Change_topG[List_Mate_Change_topGT[1].index(i)] for i in SecondQPN if i in List_Mate_Change_topGT[1]]
            SecondRowT = list(map(list, zip(*SecondRow)))
            List_Mate_Change_topGT = list(map(list, zip(*List_Mate_Change_topG)))
            SecondRowT[1] = SecondQPN
            SecondRowT[3] = List_Mate_Change_topGT[3] 
            SecondRowT[4] = List_Mate_Change_topGT[4] 
            SecondRow = list(map(list, zip(*SecondRowT)))
            PDM_top = PDM_top + SecondRow
            PDM_top = sorted(PDM_top, key = lambda a: (a[Col_Item],a[Col_PN]))
            PDM_topT = list(map(list, zip(*PDM_top)))
    
            #--- Pick out Botside's 2nd Source and fill in group & usage by rule ---
            List_Mate_ChangeAll_bot = list(set(List_Mate_ChangeAll) & (set(PDM_botT[1])))
            SecondRow = []
            SecondQPN = []
            List_Mate_Change_botG = []
            for j in List_Mate_ChangeAll_bot:
                Group = GroupwListAllT[0][GroupwListAllT[1].index(j)]
                Count = GroupwListT[0].count(Group)
                Second = [PDM_bot[PDM_botT[Col_PN].index(j)].copy()]*(Count)
                SecondRow = SecondRow + Second
                SecondQPN = SecondQPN + GroupwListT[1][GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
                del(PDM_bot[PDM_botT[Col_PN].index(j)])
                PDM_botT[Col_PN].remove(j)            
                List_Mate_Change_botG = List_Mate_Change_botG + GroupwList[GroupwListT[0].index(Group):GroupwListT[0].index(Group)+Count]
            
            List_Mate_Change_botGT = list(map(list, zip(*List_Mate_Change_botG)))
            for k in List_Mate_Change_botGT[1]:
                if 'downBOM41T_QIGU' in dir() and k in downBOM41T_QIGU[0] and downBOM41_QIGU[downBOM41T_QIGU[0].index(k)][-2] is not None:
                    List_Mate_Change_botG[List_Mate_Change_botGT[1].index(k)].extend(downBOM41_QIGU[downBOM41T_QIGU[0].index(k)][-2:])
                else:
                    List_Mate_Change_botG[List_Mate_Change_botGT[1].index(k)].extend(["",""])
            GroupwList = list(map(list, zip(*GroupwListT)))        
            if 'downBOM41T_QIGU' in dir():
                GMark_bot = list(max(max(x for x in downBOM41T_QIGU[2] if x is not None),max(y for y in PDM_botT[Col_Grp])))
            else:
                GMark_bot = list(max(max(y for y in PDM_botT[Col_Grp]),"A0"))
            WtN = ord(GMark_bot[0])*10+int(GMark_bot[1])
            List_Mate_Change_botG = sorted(List_Mate_Change_botG ,key = lambda a:(a[0],a[-1],a[2]),reverse=True)
            List_Mate_Change_botGT = list(map(list, zip(*List_Mate_Change_botG)))
            for m in list(set(List_Mate_Change_botGT[0])):
                Group = List_Mate_Change_botGT[0].index(m)
                Count = List_Mate_Change_botGT[0].count(m)
                if Count == 1:
                    pass
                elif List_Mate_Change_botGT[-1][Group] == 100:
                    List_Mate_Change_botGT[-2][Group:Group+Count]=[List_Mate_Change_botGT[-2][Group]]*len(List_Mate_Change_botGT[-2][Group:Group+Count])
                    List_Mate_Change_botGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Change_botGT[-1][Group+1:Group+Count])
                elif List_Mate_Change_botGT[-1][Group] == 0:
                    List_Mate_Change_botGT[-2][Group:Group+Count]=[List_Mate_Change_botGT[-2][Group]]*len(List_Mate_Change_botGT[-2][Group:Group+Count])
                    List_Mate_Change_botGT[-1][Group+1]=100
                    List_Mate_Change_botGT[-1][Group+2:Group+Count]=[0]*len(List_Mate_Change_botGT[-1][Group+2:Group+Count])
                elif List_Mate_Change_botGT[-1][Group] == "":
                    WtN = WtN+1
                    if WtN%10 == 0:
                        WtN = WtN + 1
                    else:
                        pass
                    NtW = chr(WtN//10)+str(WtN%10)
                    List_Mate_Change_botGT[-2][Group:Group+Count]=[NtW]*len(List_Mate_Change_botGT[-2][Group:Group+Count])
                    List_Mate_Change_botGT[-1][Group]=100
                    List_Mate_Change_botGT[-1][Group+1:Group+Count]=[0]*len(List_Mate_Change_botGT[-1][Group+1:Group+Count])             
            List_Mate_Change_botG = list(map(list, zip(*List_Mate_Change_botGT)))    
            List_Mate_Change_botG = [List_Mate_Change_botG[List_Mate_Change_botGT[1].index(i)] for i in SecondQPN if i in List_Mate_Change_botGT[1]]
            SecondRowT = list(map(list, zip(*SecondRow)))
            List_Mate_Change_botGT = list(map(list, zip(*List_Mate_Change_botG)))
            SecondRowT[1] = SecondQPN
            SecondRowT[3] = List_Mate_Change_botGT[3] 
            SecondRowT[4] = List_Mate_Change_botGT[4]
            SecondRow = list(map(list, zip(*SecondRowT)))
            PDM_bot = PDM_bot + SecondRow
            PDM_bot = sorted(PDM_bot, key = lambda a: (a[Col_Item],a[Col_PN]))
            PDM_botT = list(map(list, zip(*PDM_bot)))
        
        """
        #--- For Location Add Material ---
        if Modify_Loca_Add == []:
            pass
        else:
        
        #--- For Location Del Material ---
        if Modify_Loca_Del == []:
            pass
        else:
        
        #--- For Location Change Material ---
        if Modify_Loca_Change == []:
            pass
        else:
        """
    elif os.path.exists(ModifyList) is False:
        pass
    
    #--- fill in item number ---
    if os.path.exists(downBOM) is False:
        #--- filling in item number in botton side ---
        PDM_bot_wo2nd = []
        PDM_bot_w2nd = []
        item = 10
        for n in range(len(PDM_bot)):
           if PDM_bot[n][Col_Item] == "10":
               PDM_bot[n][Col_Item] = 10
               x = PDM_bot[n]
               PDM_bot_wo2nd.append(x)
           elif PDM_bot[n][Col_Usage] is not 0:
               item = item + 10
               PDM_bot[n][Col_Item] = item
               x = PDM_bot[n]
               PDM_bot_wo2nd.append(x)
           else:
               x = PDM_bot[n]
               PDM_bot_w2nd.append(x)
        PDM_botT_wo2nd = list(map(list, zip(*PDM_bot_wo2nd)))
        PDM_botT_w2nd = list(map(list, zip(*PDM_bot_w2nd)))
        for n in range(len(PDM_botT_w2nd[Col_PN])):
            PDM_bot_w2nd[n][Col_Item] = PDM_bot_wo2nd[PDM_botT_wo2nd[Col_Grp].index(PDM_bot_w2nd[n][Col_Grp])][Col_Item]
        PDM_bot = PDM_bot_wo2nd + PDM_bot_w2nd
        PDM_bot = sorted(PDM_bot, key = lambda a: a[Col_Item])
        
        #--- filling in item number in top side ---
        PDM_top_wo2nd = []
        PDM_top_w2nd = []
        item = 10
        for n in range(len(PDM_top)):
           if PDM_top[n][Col_Item] == "10":
               PDM_top[n][Col_Item] = 10
               x = PDM_top[n]
               PDM_top_wo2nd.append(x)
           elif PDM_top[n][Col_Usage] is not 0:
               item = item + 10
               PDM_top[n][Col_Item] = item
               x = PDM_top[n]
               PDM_top_wo2nd.append(x)
           else:
               x = PDM_top[n]
               PDM_top_w2nd.append(x)
        PDM_topT_wo2nd = list(map(list, zip(*PDM_top_wo2nd)))
        PDM_topT_w2nd = list(map(list, zip(*PDM_top_w2nd)))
        for n in range(len(PDM_topT_w2nd[Col_PN])):
            PDM_top_w2nd[n][Col_Item] = PDM_top_wo2nd[PDM_topT_wo2nd[Col_Grp].index(PDM_top_w2nd[n][Col_Grp])][Col_Item]
        PDM_top = PDM_top_wo2nd + PDM_top_w2nd
        PDM_top = sorted(PDM_top, key = lambda a: a[Col_Item])
        
        PDM_botT = list(map(list, zip(*PDM_bot)))
        PDM_topT = list(map(list, zip(*PDM_top)))
        
    #--- mapping the item number from downBOM to BOM ---
    elif os.path.exists(downBOM) is True:
        """
        pass
        """
        PDM_bot_I = [PDM_bot[i] for i in range(len(PDM_bot)) if PDM_bot[i][Col_PN] in downBOM41T_QIGU[0]]
        PDM_bot_D = [PDM_bot[i] for i in range(len(PDM_bot)) if PDM_bot[i][Col_PN] not in downBOM41T_QIGU[0]]
        PDM_top_I = [PDM_top[i] for i in range(len(PDM_top)) if PDM_top[i][Col_PN] in downBOM51T_QIGU[0]]
        PDM_top_D = [PDM_top[i] for i in range(len(PDM_top)) if PDM_top[i][Col_PN] not in downBOM51T_QIGU[0]]    
        downBOM41_QIGU_I = [downBOM41_QIGU[i] for i in range(len(downBOM41_QIGU)) if downBOM41_QIGU[i][0] in PDM_botT[Col_PN]]
        downBOM51_QIGU_I = [downBOM51_QIGU[i] for i in range(len(downBOM51_QIGU)) if downBOM51_QIGU[i][0] in PDM_topT[Col_PN]]
        PDM_bot_I = sorted(PDM_bot_I, key = lambda a: a[Col_PN])
        PDM_top_I = sorted(PDM_top_I, key = lambda a: a[Col_PN])
        downBOM41_QIGU_I = sorted(downBOM41_QIGU_I, key = lambda a: a[0])
        downBOM51_QIGU_I = sorted(downBOM51_QIGU_I, key = lambda a: a[0])
        PDM_botT_I = list(map(list, zip(*PDM_bot_I)))
        PDM_topT_I = list(map(list, zip(*PDM_top_I)))
        downBOM41T_QIGU_I = list(map(list, zip(*downBOM41_QIGU_I)))
        downBOM51T_QIGU_I = list(map(list, zip(*downBOM51_QIGU_I)))
        PDM_botT_I[Col_Item] = downBOM41T_QIGU_I[1]
        PDM_topT_I[Col_Item] = downBOM51T_QIGU_I[1]
        PDM_bot_I = list(map(list, zip(*PDM_botT_I)))
        PDM_top_I = list(map(list, zip(*PDM_topT_I)))
        PDM_botT_D = list(map(list, zip(*PDM_bot_D)))
        PDM_topT_D = list(map(list, zip(*PDM_top_D)))
        
        #--- fill the new item number in botton side ---    
        PDM_bot_DwoItem = []
        PDM_bot_DwItem = []
        PDM_bot_DwTemp = []
        for n in range(len(PDM_bot_D)):
            if PDM_bot_D[n][Col_Item] == "10":
               PDM_bot_D[n][Col_Item] = 10
               x = PDM_bot_D[n]
               PDM_bot_DwItem.append(x)
            elif PDM_bot_D[n][Col_Usage] is "":
               x = PDM_bot_D[n]
               PDM_bot_DwoItem.append(x)
            else:
                if PDM_bot_D[n][Col_Grp] in PDM_botT_I[Col_Grp]:
                    PDM_bot_D[n][Col_Item] = PDM_botT_I[Col_Item][PDM_botT_I[Col_Grp].index(PDM_bot_D[n][Col_Grp])]
                    y = PDM_bot_D[n]
                    PDM_bot_DwItem.append(y)
                elif PDM_bot_D[n][Col_Usage] is 100:
                    x = PDM_bot_D[n]
                    PDM_bot_DwoItem.append(x)
                else:
                    z = PDM_bot_D[n]
                    PDM_bot_DwTemp.append(z)
        
        PDM_bot_DwoItem = sorted(PDM_bot_DwoItem, key = lambda a: a[Col_PN])
        PDM_bot_DwoItemT = list(map(list, zip(*PDM_bot_DwoItem)))    
        item = max(i for i in downBOM41T_QIGU[1] if type(i) is int)    
        for n in range(len(PDM_bot_DwoItemT[Col_PN])):
            item = item + 10
            PDM_bot_DwoItem[n][Col_Item] = item
            x = PDM_bot_DwoItem[n]
            PDM_bot_DwItem.append(x)
        PDM_bot_DwItemT = list(map(list, zip(*PDM_bot_DwItem)))       
        for n in range(len(PDM_bot_DwTemp)):
            PDM_bot_DwTemp[n][Col_Item] = PDM_bot_DwItemT[Col_Item][PDM_bot_DwItemT[Col_Grp].index(PDM_bot_DwTemp[n][Col_Grp])]
        
        PDM_bot = PDM_bot_I + PDM_bot_DwItem + PDM_bot_DwTemp
        PDM_bot = sorted(PDM_bot, key = lambda a: a[Col_Item])
        PDM_botT = list(map(list, zip(*PDM_bot)))
    
        #--- fill the new item number in top side ---    
        PDM_top_DwoItem = []
        PDM_top_DwItem = []
        PDM_top_DwTemp = []
        for n in range(len(PDM_top_D)):
            if PDM_top_D[n][Col_Item] == "10":
               PDM_top_D[n][Col_Item] = 10
               x = PDM_top_D[n]
               PDM_top_DwItem.append(x)
            elif PDM_top_D[n][Col_Usage] is "":
               x = PDM_top_D[n]
               PDM_top_DwoItem.append(x)
            else:
                if PDM_top_D[n][Col_Grp] in PDM_topT_I[Col_Grp]:
                    PDM_top_D[n][Col_Item] = PDM_topT_I[Col_Item][PDM_topT_I[Col_Grp].index(PDM_top_D[n][Col_Grp])]
                    y = PDM_top_D[n]
                    PDM_top_DwItem.append(y)
                elif PDM_top_D[n][Col_Usage] is 100:
                    x = PDM_top_D[n]
                    PDM_top_DwoItem.append(x)
                else:
                    z = PDM_top_D[n]
                    PDM_top_DwTemp.append(z)
        
        PDM_top_DwoItem = sorted(PDM_top_DwoItem, key = lambda a: a[Col_PN])
        PDM_top_DwoItemT = list(map(list, zip(*PDM_top_DwoItem)))    
        item = max(i for i in downBOM51T_QIGU[1] if type(i) is int)    
        for n in range(len(PDM_top_DwoItemT[Col_PN])):
            item = item + 10
            PDM_top_DwoItem[n][Col_Item] = item
            x = PDM_top_DwoItem[n]
            PDM_top_DwItem.append(x)
        PDM_top_DwItemT = list(map(list, zip(*PDM_top_DwItem)))       
        for n in range(len(PDM_top_DwTemp)):
            PDM_top_DwTemp[n][Col_Item] = PDM_top_DwItemT[Col_Item][PDM_top_DwItemT[Col_Grp].index(PDM_top_DwTemp[n][Col_Grp])]
        
        PDM_top = PDM_top_I + PDM_top_DwItem + PDM_top_DwTemp
        PDM_top = sorted(PDM_top, key = lambda a: a[Col_Item])
        PDM_topT = list(map(list, zip(*PDM_top))) 
    
    """ error check
    BomFile = "G7C WHLU R17M 2G HYNIX 0416"
    BomAddress = "D:/Project/G7CD/DB1/DSN/2G/"+BomFile+".BOM"
    Bom = open(BomAddress,"r").readlines()
    BomList = [Bom[i].split("\t") for i in range(len(Bom))]
    Aaa = []
    Aaa = [BomList[i] for i in range(len(BomList)) if len(BomList[i]) != 5 ]
    Ccc = []
    Ccc = [BomList[i] for i in range(len(BomList)) if len(BomList[i]) == 5 ]
    Bbb = []
    Bbb = [Ccc[i] for i in range(len(Ccc)) if len(Ccc[i][-1]) !=4]
    BomListTT = list(map(list, zip(*BomList)))
    """
    
    PDM_bot.insert(0,Head)
    PDM_top.insert(0,Head)
    
    wb_bot = excel.Workbook()
    ws_bot = wb_bot.active
    ws_bot.title = "pdm_bot"
    for row in range(len(PDM_bot)):
        for col in range(len(PDM_bot[row])):
            ws_bot.cell(row=row+1, column=col+1).value = PDM_bot[row][col]
    wb_bot.save(QPN_41+".xlsx")         
            
    wb_top = excel.Workbook()
    ws_top = wb_top.active
    ws_top.title = "pdm_top"
    for row in range(len(PDM_top)):
        for col in range(len(PDM_top[row])):
            ws_top.cell(row=row+1, column=col+1).value = PDM_top[row][col]
    wb_top.save(QPN_51+".xlsx")     
    TimeEnd = time.time()
    DeltaTime = TimeEnd - TimeStart
    return DeltaTime